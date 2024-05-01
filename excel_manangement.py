import os.path
import xlrd
import openpyxl
import logging
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from common_fn import exception_log

logger = logging.getLogger('my_logger')


# to store all the key values in a dict from a json,
# used in get_all_keys_values function
list_of_dict = list()


class XlsFile:
    def __init__(self, xls):
        self.xls = xls

    def __enter__(self):
        self.book = xlrd.open_workbook(self.xls)
        return self.book

    def __exit__(self, *args):
        self.book.release_resources()


class OpenXlsxFile:
    def __init__(self, xlsx):
        self.xlsx = xlsx

    def __enter__(self):
        self.workbook = openpyxl.Workbook()
        return self.workbook

    def __exit__(self, *args):
        self.workbook.save(self.xlsx)
        self.workbook.close()


class LoadXlsxFile:
    def __init__(self, xlsx: str,
                 read_only: bool = False,
                 keep_vba: bool = False,
                 data_only: bool = False,
                 keep_links: bool = True,
                 rich_text: bool = False):
        self.xlsx = xlsx
        self.read_only = read_only
        self.keep_vba = keep_vba
        self.data_only = data_only
        self.keep_links = keep_links
        self.rich_text = rich_text

    def __enter__(self):
        self.workbook = openpyxl.load_workbook(self.xlsx, self.read_only, self.keep_vba,
                                               self.data_only, self.keep_links, self.rich_text)
        return self.workbook

    def __exit__(self, *args):
        self.workbook.save(self.xlsx)
        self.workbook.close()


def read_rows(excel_file_path: str, start_row_index: int | None = None,
              end_row_index: int | None = None, sheet_index: int = 0) -> list | bool:
    if start_row_index == 0:
        # print('Start row index cannot be 0.')
        logger.warning('Start row index cannot be 0.')
        return False

    try:
        with XlsFile(excel_file_path) as book:
            sh = book.sheet_by_index(sheet_index)

            rows_list = []

            # get all rows from start
            if start_row_index is None and end_row_index is None:
                rows_list = [sh.row(row) for row in range(2, sh.nrows + 1)]

            # get rows from start to desired index number
            elif start_row_index is None and end_row_index is not None:
                rows_list = [sh.row(row) for row in range(2, end_row_index + 1)]

            # get rows from desired index number to end
            elif start_row_index is not None and end_row_index is None:
                rows_list = [sh.row(row) for row in range(start_row_index - 1, sh.nrows)]

            # get rows from given start to given end index number
            elif start_row_index is not None and end_row_index is not None:
                rows_list = [sh.row(row) for row in range(start_row_index-1, end_row_index)]

        return rows_list

    except Exception as e:
        exception_log(e)


def write_in_new_xlsx(file_path: str, sheet_name, data_to_write: dict, column_index: int = 1):
    try:
        with OpenXlsxFile(file_path) as wb:
            wb.create_sheet(title=str(sheet_name), index=0)
            wb.remove(wb.worksheets[-1])

            sheet = wb.active
            list_of_dict.clear()  # clear old data if there is any
            data_to_write = get_all_keys_values(data_to_write)

            for i, data_dict in enumerate(data_to_write):
                for k, v in data_dict.items():
                    cell_ref = sheet.cell(row=i + 1, column=column_index)
                    cell_ref.value = str(k)

                    cell_ref = sheet.cell(row=i + 1, column=column_index + 1)
                    cell_ref.value = str(v)

    except Exception as e:
        exception_log(e)


def write_in_existing_xlsx(file_path: str, sheet_name, data_to_write: dict, column_index: int = 1):
    try:
        with LoadXlsxFile(file_path) as wb:
            if str(sheet_name) not in wb.sheetnames:
                wb.create_sheet(title=str(sheet_name))

            wb.active = wb[f'{sheet_name}']
            sheet = wb.active
            list_of_dict.clear()  # clear old data if there is any
            data_to_write = get_all_keys_values(data_to_write)

            for i, data_dict in enumerate(data_to_write):
                for k, v in data_dict.items():
                    cell_ref = sheet.cell(row=i + 1, column=column_index)
                    cell_ref.value = str(k)

                    cell_ref = sheet.cell(row=i + 1, column=column_index + 1)
                    cell_ref.value = str(v)
                    cell_ref.value = str(v)

    except Exception as e:
        exception_log(e)


def write_in_xl(file_path: str, sheet_name, data_to_write: dict, column_index: int = 1):
    if os.path.isfile(file_path):
        write_in_existing_xlsx(file_path, sheet_name, data_to_write, column_index)
    else:
        write_in_new_xlsx(file_path, sheet_name, data_to_write, column_index)


def compare_data(file_path: str, sheet_name, ref_column: int, compare_to_column: int, compare_in_column: int):
    try:
        with LoadXlsxFile(file_path) as wb:
            wb.active = wb[str(f'{sheet_name}')]
            sheet = wb.active

            i = 1
            while True:
                if sheet.cell(row=i, column=ref_column).value is not None or \
                        sheet.cell(row=i, column=compare_to_column).value is not None:

                    cell_ref = sheet.cell(row=i, column=compare_in_column)

                    cell_ref.value = \
                        (f'=IF({cell_ref.offset(row=i, column=ref_column - compare_in_column).column_letter}{i}='
                         f'{cell_ref.offset(row=i, column=compare_to_column - compare_in_column).column_letter}{i}, '
                         f'"Pass", "Fail")')

                else:
                    break

                i += 1

        set_cond_formatting(file_path, sheet_name, f'{cell_ref.column_letter}{1}:{cell_ref.column_letter}{i - 1}')

    except Exception as e:
        exception_log(e)


def set_cond_formatting(file_path: str, sheet_name, cell_range):
    try:
        with LoadXlsxFile(file_path) as wb:
            wb.active = wb[str(f'{sheet_name}')]
            sheet = wb.active

            # creating green_fill and red_fill
            green_fill = PatternFill(start_color='75F94D', end_color='75F94D', fill_type='solid')
            red_fill = PatternFill(start_color='EB1212', end_color='EB1212', fill_type='solid')

            sheet.conditional_formatting.add(f'{cell_range}', FormulaRule(
                formula=[f'NOT(ISERROR(SEARCH("Pass",{cell_range[:cell_range.find(':')]})))'],
                stopIfTrue=True, fill=green_fill))

            sheet.conditional_formatting.add(f'{cell_range}', FormulaRule(
                formula=[f'NOT(ISERROR(SEARCH("Fail",{cell_range[:cell_range.find(':')]})))'],
                stopIfTrue=True, fill=red_fill))

    except Exception as e:
        exception_log(e)


def get_all_keys_values(data: dict) -> list[{str, str}]:
    try:
        temp_list = list()
        for k, v in data.items():
            if type(v) is dict:
                get_all_keys_values(v)

            elif type(v) is list:
                for i in v:
                    if type(i) is dict:
                        temp_list.append(k)
                        get_all_keys_values(i)
                    elif type(i) is list:
                        # print(k, v)
                        list_of_dict.append({k: v})
                else:
                    if k not in temp_list:
                        # print(k, v)
                        list_of_dict.append({k: v})
            else:
                # print(k, v)
                list_of_dict.append({k: v})

        return list_of_dict

    except Exception as e:
        exception_log(e)


if __name__ == '__main__':
    a = {"SupplyType": 0, "ServicePointNo": "SAM000012597", "DeviceNo": "SAM000012597", "ScheduledTasks": [
        {"ScheduleTime": {"StartTime": {"Hour": 0, "Minute": 0, "Second": 0}, "Frequency": 39}, "DataType": 0},
        {"ScheduleTime": {"StartTime": {"Hour": 0, "Minute": 0, "Second": 0}, "Frequency": 39}, "DataType": 1},
        {"ScheduleTime": {"StartTime": {"Hour": 0, "Minute": 0, "Second": 0}, "Frequency": 34}, "DataType": 2},
        {"ScheduleTime": {"StartTime": {"Hour": 0, "Minute": 0, "Second": 0}, "Frequency": 39}, "DataType": 3,
         "ProfileParameters": [0, 2]},
        {"ScheduleTime": {"StartTime": {"Hour": 0, "Minute": 0, "Second": 0}, "Frequency": 39}, "DataType": 4,
         "ProfileParameters": [6]},
        {"ScheduleTime": {"StartTime": {"Hour": 0, "Minute": 0, "Second": 0}, "Frequency": 39}, "DataType": 5,
         "EventLogs": [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]}]}

    write_in_new_xlsx(r'.\supporting\my_sheet.xlsx', 1, a, 1)
    write_in_existing_xlsx(r'.\supporting\my_sheet.xlsx', 1, a, 3)
    compare_data(r'.\supporting\my_sheet.xlsx', 1, 1, 3, 5)
    compare_data(r'.\supporting\my_sheet.xlsx', 1, 2, 4, 6)
